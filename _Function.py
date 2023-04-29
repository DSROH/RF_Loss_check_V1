import os, sys, subprocess
from operator import sub
import matplotlib.pyplot as plt
import pandas as pd
import re

plt.rc("font", family="Malgun Gothic")
plt.rc("axes", unicode_minus=False)

from matplotlib.backends.backend_pdf import PdfPages
import tkinter as tk
from tkinter import filedialog
import tkinter.messagebox as msgbox

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.numbers import builtin_format_code
from _RF_loss_Spec import Freq_list_129


def save_multi_image(f_name):
    pp = PdfPages(f_name)
    fig_nums = plt.get_fignums()
    figs = [plt.figure(n) for n in fig_nums]
    for fig in figs:
        fig.savefig(pp, format="pdf")
    pp.close()


def open_file(filename):
    if sys.platform == "win32":
        os.startfile(filename)
    else:
        opener = "open" if sys.platform == "darwin" else "xdg-open"
        subprocess.call([opener, filename])


def get_aspect(ax):
    # Total figure size
    figW, figH = ax.get_figure().get_size_inches()
    # Axis size on figure
    _, _, w, h = ax.get_position().bounds
    # Ratio of display units
    disp_ratio = (figH * h) / (figW * w)
    # Ratio of data units
    # Negative over negative because of the order of subtraction
    data_ratio = sub(*ax.get_ylim()) / sub(*ax.get_xlim())

    return disp_ratio / data_ratio


def add_file(path, list_file):
    list_file.delete(0, tk.END)
    if path == "Daseul":
        init_dir = "C:\DGS\LOGS"
    elif path == "Path":
        init_dir = "C:\DGS\LOGS\PATHLOSS"
    filename = filedialog.askopenfilenames(
        initialdir=init_dir,
        title="Select file",
        filetypes=(("All fiels", "*.*"), ("Excel files", "*.xlsx")),
    )

    for file in filename:
        list_file.insert(tk.END, file)


def browse_lossfile(path_lossfile, Selected_lossfile):
    spc_file_name = filedialog.askopenfilename(
        title="SPC 파일을 선택하세요",
        filetypes=(("atten_table", "*.txt"), ("모든 파일", "*.*")),
        initialdir=r"D:\\DATA\\Project_DATA\\@_S23_FE\\TOOL\\2_MTM Calibration",
    )
    # print(folder_selected)
    path_lossfile.delete(0, tk.END)
    path_lossfile.insert(0, spc_file_name)
    Selected_lossfile.set(True)


def Common_save_Excel(filename, tab1):
    # Save Data to Excel
    Tabname = filename.replace("Excel_", "")
    Tabname = f"{os.path.splitext(Tabname)[0]}"
    with pd.ExcelWriter(filename) as writer:
        tab1.to_excel(writer, sheet_name=f"{Tabname}_Mean")


def WB_Format(filename, i, j, k):
    font_style = Font(
        name="Calibri",
        size=10,
        bold=False,
        italic=False,
        vertAlign=None,  # 첨자
        underline="none",  # 밑줄
        strike=False,  # 취소선
        color="00000000",  # 블랙, # 00FF0000 Red, # 000000FF Blue
    )
    wb = load_workbook(filename)
    ws = wb.sheetnames
    for sheet in ws:
        col_max = wb[sheet].max_column
        row_max = wb[sheet].max_row
        for row_c in range(i, row_max + 1, 1):
            for col_c in range(j, col_max + 1, 1):
                wb[sheet].cell(row=row_c, column=col_c).font = font_style
                wb[sheet].cell(row=row_c, column=col_c).alignment = Alignment(horizontal="right")
                # wb[sheet].cell(row=row_c, column=col_c).number_format = "#,##0.0"
                wb[sheet].cell(row=row_c, column=col_c).number_format = builtin_format_code(k)
    wb.save(filename)


def get_data(filename, Result_var):
    df_BtoB1 = pd.DataFrame()

    if filename:
        for FileNumber, file in enumerate(filename):
            fname_only = os.path.basename(file).split(".")[0]
            # Import Data
            my_cols = [str(i) for i in range(10)]  # create some col names
            df_Data = pd.read_csv(file, sep="\t|,", names=my_cols, header=None, engine="python")
            Current_Type = df_Data[df_Data["0"].str.contains("Current Cable Type", na=False)]
            Current_Type = Current_Type["0"].str.split(":", expand=True).iloc[0, 1].strip()
            df_Test = df_Data.index[(df_Data["0"] == "#TEST")].to_list()

            if df_Data["0"].str.contains("SVC", na=False).any():
                df_loss = df_Data[df_Data["0"].str.contains("RF Cable Type BtoB : ", na=False)].iloc[:, :1]
                Type_SVC = True
            elif Current_Type == "BtoB":
                df_loss = df_Data[df_Data["0"].str.contains("RF Cable Type BtoB : ", na=False)].iloc[:, :1]
                Type_SVC = False
            else:
                df_loss = df_Data[df_Data["0"].str.contains("RF Cable Type : ", na=False)].iloc[:, :1]
                Type_SVC = False

            Jig_list = df_Data[df_Data["0"].str.contains("JIG :", na=False)].iloc[:, :1]
            Result = df_Data[df_Data["0"].str.contains("RESULT :", na=False)].iloc[:, :1]
            lineip_list = df_Data[df_Data["0"].str.contains("RDM_LOT :", na=False)].iloc[:, :1]

            for Count, index in enumerate(df_Test):
                # for index in BtoB1:
                LossCal_Result = Result.iloc[Count].to_list()[0].split(":")[1]
                lineip = lineip_list.iloc[Count].to_list()[0].split(":")[1]
                lineip = lineip.split("_")[0].strip()

                if Result_var:
                    if Type_SVC:
                        BtoB_Type = "SVC"
                    else:
                        BtoB_Type = "BtoB"

                    Type = int(df_loss.iloc[Count].to_list()[0].split(":")[1].strip())
                    Jig = Jig_list.iloc[Count].to_list()[0].strip()

                    if (Type == 18) or (Type == 19):
                        Size = 98
                    elif Type == 7:
                        Size = 98
                    elif Type == 62:
                        Size = 129
                    else:
                        Size = 58

                    Loss_Start = index + 3
                    Loss_Stop = Loss_Start + Size
                    df_BtoB_1st = df_Data.iloc[Loss_Start:Loss_Stop, :4]
                    BtoB1st_Value = df_BtoB_1st.iloc[:, 1:2].reset_index(drop=True)
                    BtoB1st_Value = BtoB1st_Value.astype(float)
                    BtoB1st_Value.columns = [f"{fname_only}_IP_{lineip}_{Jig}"]
                    BtoB1st_Item = df_BtoB_1st["0"].str.split(" ", expand=True).reset_index(drop=True)

                    if BtoB1st_Item[0].str.contains("SVC", na=False).any():
                        BtoB1st_Item.columns = ["SVC", "Meas", "BtoB_No", "Path", "Frequency"]
                    elif Current_Type == "BtoB":
                        BtoB1st_Item.columns = ["Meas", "BtoB_No", "Path", "Frequency"]
                    else:
                        BtoB1st_Item.columns = ["Meas", "Path", "Frequency"]

                    BtoB1st_Item["Frequency"] = BtoB1st_Item["Frequency"].str.extract(r"(\d+)").astype(int)
                    BtoB1st_Value.index = BtoB1st_Item["Frequency"]
                    df_BtoB1 = pd.concat([df_BtoB1, BtoB1st_Value], axis=1)

                else:
                    if LossCal_Result == "FAIL":
                        msgbox.showwarning("Warning", f"Losscal Result : Fail\nOr\nCheck 'Include Failed log' Button")
                        plt.close()
                        return
                    else:
                        if Type_SVC:
                            BtoB_Type = "SVC"
                        else:
                            BtoB_Type = "BtoB"

                        Type = int(df_loss.iloc[Count].to_list()[0].split(":")[1].strip())
                        Jig = Jig_list.iloc[Count].to_list()[0].strip()

                        if (Type == 18) or (Type == 19):
                            Size = 98
                        elif Type == 7:
                            Size = 98
                        elif Type == 62:
                            Size = 129
                        else:
                            Size = 58

                        Loss_Start = index + 3
                        Loss_Stop = Loss_Start + Size
                        df_BtoB_1st = df_Data.iloc[Loss_Start:Loss_Stop, :4]
                        BtoB1st_Value = df_BtoB_1st.iloc[:, 1:2].reset_index(drop=True)
                        BtoB1st_Value = BtoB1st_Value.astype(float)
                        BtoB1st_Value.columns = [f"{fname_only}_IP_{lineip}_{Jig}"]
                        BtoB1st_Item = df_BtoB_1st["0"].str.split(" ", expand=True).reset_index(drop=True)

                        if BtoB1st_Item[0].str.contains("SVC", na=False).any():
                            BtoB1st_Item.columns = ["SVC", "Meas", "BtoB_No", "Path", "Frequency"]
                        elif Current_Type == "BtoB":
                            BtoB1st_Item.columns = ["Meas", "BtoB_No", "Path", "Frequency"]
                        else:
                            BtoB1st_Item.columns = ["Meas", "Path", "Frequency"]

                        BtoB1st_Item["Frequency"] = BtoB1st_Item["Frequency"].str.extract(r"(\d+)").astype(int)
                        BtoB1st_Value.index = BtoB1st_Item["Frequency"]
                        df_BtoB1 = pd.concat([df_BtoB1, BtoB1st_Value], axis=1)

    df_meas_mean = round(df_BtoB1.groupby(["Frequency"], sort=False).mean(), 2)
    df_meas_mean["Average"] = round(df_meas_mean.mean(axis=1), 2)
    df_meas_mean["Max"] = round(df_meas_mean.max(axis=1), 2)
    df_meas_mean["Min"] = round(df_meas_mean.min(axis=1), 2)

    fname = "Excel_Atten_table.xlsx"
    Common_save_Excel(fname, df_meas_mean)
    WB_Format(fname, 2, 2, 4)

    return df_meas_mean["Average"], Size


def Set_atten_file_format(path_lossfile, index_count):
    target_word = f"MaxIndex="
    new_text_content = ""

    with open(path_lossfile, "r", encoding="utf-8") as file:
        data_lines = file.readlines()
    file.close()

    for index, line in enumerate(data_lines):
        if target_word in line:
            New_String = re.split("=|\n", line)
            New_String = [v for v in New_String if v]
            Freq = int(re.sub(r"[^0-9]", "", New_String[1]))
            new_text_content += line
            break
        else:
            new_text_content += line

    if Freq < index_count:  # 98 포인트
        for index in range(1, index_count + 1, 1):
            if index < 100:
                replaced_Str = f"Frequency_{index:02d}={Freq_list_129[index-1]}\n"
                new_text_content += replaced_Str
                replaced_Str = f"RFLoss_{index:02d}=-999\n"
                new_text_content += replaced_Str
            else:
                replaced_Str = f"Frequency_{index:03d}={Freq_list_129[index-1]}\n"
                new_text_content += replaced_Str
                replaced_Str = f"RFLoss_{index:03d}=-999\n"
                new_text_content += replaced_Str

        with open(path_lossfile, "w", encoding="utf-8") as f:
            f.writelines(new_text_content)
        f.close()
    else:  # 129포인트
        pass


def Chng_loss(path_lossfile, df_meas, index_count):
    target_word = f"MaxIndex="
    new_text_content = ""
    Check = False

    with open(path_lossfile, "r", encoding="utf-8") as file:
        data_lines = file.readlines()
    file.close()

    for index, line in enumerate(data_lines):
        if target_word in line:
            Old_String = line
            Check = True
            New_String = re.split("=|\n", line)
            New_String = [v for v in New_String if v]
            New_String[1] = str(index_count)
            New_String = "=".join(New_String) + "\n"
            replaced_Str = line.replace(Old_String, New_String)
            new_text_content += replaced_Str
        elif Check & line.startswith("Frequency_"):
            New_String = Old_String = line
            New_String = re.split("=|\n", line)
            New_String = [v for v in New_String if v]
            freq = int(re.sub(r"[^0-9]", "", New_String[1]))
            new_text_content += line
        elif Check & line.startswith("RFLoss_"):
            New_String = Old_String = line
            New_String = re.split("=|\n", line)
            New_String = [v for v in New_String if v]
            New_String[1] = df_meas[freq]
            New_String = [str(v) for v in New_String if v]
            New_String = "=".join(New_String) + "\n"
            replaced_Str = line.replace(Old_String, New_String)
            new_text_content += replaced_Str
        elif line.startswith("[Measured_CableLoss_02]"):
            Check = False
            new_text_content += line
        else:
            new_text_content += line

    with open(path_lossfile, "w", encoding="utf-8") as f:
        f.writelines(new_text_content)
    f.close()


def transf_to_attentable(filename, path_lossfile, Result_var, Selected_lossfile):
    if Selected_lossfile:
        try:
            df_meas, index_count = get_data(filename, Result_var)
            Set_atten_file_format(path_lossfile, index_count)
            Chng_loss(path_lossfile, df_meas, index_count)
            msgbox.showinfo("info", f"Done")
        except Exception as e:
            msgbox.showwarning("warning", e)
    else:
        msgbox.showwarning("warning", f"Select atten_table.txt")
